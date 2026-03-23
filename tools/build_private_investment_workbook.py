from __future__ import annotations

import os
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
VENDOR = ROOT / ".vendor"
if str(VENDOR) not in sys.path:
    sys.path.insert(0, str(VENDOR))

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


OUTPUT_FILE = ROOT / "Family_Office_Private_Investment_Tracker.xlsx"
ENTITIES = ["Beaman Ventures", "Lee Beaman", "Kat Trust", "Nat Trust"]
STATUSES = ["Active", "Realized", "Partially Realized", "Written Off"]
TRANSACTION_TYPES = [
    "Capital Call",
    "Distribution",
    "Dividend",
    "Fee",
    "Return of Capital",
    "Partial Exit",
    "Full Exit",
    "Write-off adjustment",
]
VALUATION_METHODS = [
    "Cost",
    "EBITDA multiple",
    "Revenue multiple",
    "EBIT multiple",
    "NAV / asset-based",
    "Mark-to-market",
    "Manual override",
]
MARK_TYPES = [
    "Cost",
    "Priced Round",
    "Third-Party Transaction",
    "Exit Process",
    "Write-down",
    "Other",
]
SECTORS = [
    "Application Software",
    "Healthcare Services",
    "Industrial Technology",
    "Private Credit",
    "Consumer",
    "Life Sciences",
    "Energy",
    "Fintech",
]
ASSET_TYPES = [
    "Venture Equity",
    "Growth Equity",
    "Private Equity",
    "Private Credit",
    "Fund",
]

COLOR_DARK_BLUE = "1F4E79"
COLOR_LIGHT_BLUE = "D9EAF7"
COLOR_YELLOW = "FFF2CC"
COLOR_FORMULA = "F4F6F8"
COLOR_GREEN = "E2F0D9"
COLOR_BORDER = "9AA6B2"
COLOR_TEXT = "1F2937"

THIN = Side(style="thin", color=COLOR_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


INVESTMENTS = [
    {
        "id": "INV-001",
        "name": "Atlas Data Platforms",
        "sector": "Application Software",
        "asset_type": "Venture Equity",
        "status": "Active",
        "date": date(2022, 3, 15),
        "sponsor": "North Peak Ventures",
        "geo": "United States",
        "notes": "Workflow automation platform serving mid-market industrial customers.",
        "official_method": "Priced Round",
        "internal_method": "Revenue multiple",
        "exit_method": "Revenue multiple",
        "board": "Y",
        "lead": "Lead",
        "commitment": 3000000,
        "bv_pct": 0.40,
        "lb_pct": 0.30,
        "kt_pct": 0.20,
        "nt_pct": 0.10,
        "own_total": 0.08,
        "entity_notes": "Cross-entity core software position.",
    },
    {
        "id": "INV-002",
        "name": "Harbor Health Partners",
        "sector": "Healthcare Services",
        "asset_type": "Growth Equity",
        "status": "Active",
        "date": date(2021, 9, 10),
        "sponsor": "Oak Hill Growth",
        "geo": "United States",
        "notes": "Provider network services platform with recurring contracted revenue.",
        "official_method": "Cost",
        "internal_method": "EBITDA multiple",
        "exit_method": "EBITDA multiple",
        "board": "N",
        "lead": "Co-invest",
        "commitment": 2500000,
        "bv_pct": 0.50,
        "lb_pct": 0.00,
        "kt_pct": 0.25,
        "nt_pct": 0.25,
        "own_total": 0.12,
        "entity_notes": "Weighted toward Beaman Ventures with trust participation.",
    },
    {
        "id": "INV-003",
        "name": "Summit Robotics",
        "sector": "Industrial Technology",
        "asset_type": "Venture Equity",
        "status": "Partially Realized",
        "date": date(2020, 5, 5),
        "sponsor": "Forge Capital",
        "geo": "United States",
        "notes": "Warehouse automation company with partial secondary liquidity realized.",
        "official_method": "Priced Round",
        "internal_method": "Revenue multiple",
        "exit_method": "EBITDA multiple",
        "board": "Y",
        "lead": "Co-invest",
        "commitment": 1200000,
        "bv_pct": 0.00,
        "lb_pct": 0.60,
        "kt_pct": 0.00,
        "nt_pct": 0.40,
        "own_total": 0.065,
        "entity_notes": "Lee / Nat co-invest with partial liquidity already received.",
    },
    {
        "id": "INV-004",
        "name": "Cedar Credit Fund II",
        "sector": "Private Credit",
        "asset_type": "Fund",
        "status": "Active",
        "date": date(2022, 1, 20),
        "sponsor": "Cedar Capital",
        "geo": "North America",
        "notes": "Senior secured direct lending fund with quarterly distributions.",
        "official_method": "Cost",
        "internal_method": "NAV / asset-based",
        "exit_method": "NAV / asset-based",
        "board": "N",
        "lead": "Co-invest",
        "commitment": 2000000,
        "bv_pct": 0.25,
        "lb_pct": 0.25,
        "kt_pct": 0.25,
        "nt_pct": 0.25,
        "own_total": 0.04,
        "entity_notes": "Evenly funded across all four entities.",
    },
    {
        "id": "INV-005",
        "name": "Meridian Consumer Holdings",
        "sector": "Consumer",
        "asset_type": "Private Equity",
        "status": "Realized",
        "date": date(2019, 7, 1),
        "sponsor": "Crestline Partners",
        "geo": "United States",
        "notes": "Fully exited consumer roll-up with strong realized gains.",
        "official_method": "Exit Process",
        "internal_method": "EBITDA multiple",
        "exit_method": "EBITDA multiple",
        "board": "N",
        "lead": "Co-invest",
        "commitment": 1500000,
        "bv_pct": 0.45,
        "lb_pct": 0.35,
        "kt_pct": 0.20,
        "nt_pct": 0.00,
        "own_total": 0.05,
        "entity_notes": "No Nat Trust participation in the original deal.",
    },
    {
        "id": "INV-006",
        "name": "Northshore BioTools",
        "sector": "Life Sciences",
        "asset_type": "Venture Equity",
        "status": "Active",
        "date": date(2023, 11, 3),
        "sponsor": "Helix Ventures",
        "geo": "United States",
        "notes": "Diagnostics tools business with strong internal growth expectations.",
        "official_method": "Cost",
        "internal_method": "Revenue multiple",
        "exit_method": "Revenue multiple",
        "board": "N",
        "lead": "Co-invest",
        "commitment": 900000,
        "bv_pct": 0.00,
        "lb_pct": 0.00,
        "kt_pct": 0.55,
        "nt_pct": 0.45,
        "own_total": 0.09,
        "entity_notes": "Trust-only exposure.",
    },
    {
        "id": "INV-007",
        "name": "Ironclad Energy Services",
        "sector": "Energy",
        "asset_type": "Private Equity",
        "status": "Written Off",
        "date": date(2018, 4, 12),
        "sponsor": "Lone Star Capital",
        "geo": "United States",
        "notes": "Legacy oilfield services investment marked to zero.",
        "official_method": "Write-down",
        "internal_method": "Manual override",
        "exit_method": "Manual override",
        "board": "N",
        "lead": "Co-invest",
        "commitment": 500000,
        "bv_pct": 0.30,
        "lb_pct": 0.20,
        "kt_pct": 0.25,
        "nt_pct": 0.25,
        "own_total": 0.03,
        "entity_notes": "Legacy position retained for historical performance tracking.",
    },
    {
        "id": "INV-008",
        "name": "Solstice Payments",
        "sector": "Fintech",
        "asset_type": "Growth Equity",
        "status": "Active",
        "date": date(2024, 6, 18),
        "sponsor": "Meridian Growth",
        "geo": "Europe",
        "notes": "Cross-border payments platform with scaling margins.",
        "official_method": "Priced Round",
        "internal_method": "Revenue multiple",
        "exit_method": "Revenue multiple",
        "board": "N",
        "lead": "Lead",
        "commitment": 1750000,
        "bv_pct": 0.35,
        "lb_pct": 0.15,
        "kt_pct": 0.25,
        "nt_pct": 0.25,
        "own_total": 0.07,
        "entity_notes": "Recent core growth holding across all entities.",
    },
]

LEDGER = [
    ["TX-001","INV-001","Atlas Data Platforms","Beaman Ventures",date(2022,3,15),"Capital Call","Initial close",400000,0,"Unrealized",None,""],
    ["TX-002","INV-001","Atlas Data Platforms","Lee Beaman",date(2022,3,15),"Capital Call","Initial close",300000,0,"Unrealized",None,""],
    ["TX-003","INV-001","Atlas Data Platforms","Kat Trust",date(2022,3,15),"Capital Call","Initial close",200000,0,"Unrealized",None,""],
    ["TX-004","INV-001","Atlas Data Platforms","Nat Trust",date(2022,3,15),"Capital Call","Initial close",100000,0,"Unrealized",None,""],
    ["TX-005","INV-001","Atlas Data Platforms","Beaman Ventures",date(2023,4,10),"Capital Call","Product expansion",120000,0,"Unrealized",None,""],
    ["TX-006","INV-001","Atlas Data Platforms","Lee Beaman",date(2023,4,10),"Capital Call","Product expansion",90000,0,"Unrealized",None,""],
    ["TX-007","INV-001","Atlas Data Platforms","Kat Trust",date(2023,4,10),"Capital Call","Product expansion",60000,0,"Unrealized",None,""],
    ["TX-008","INV-001","Atlas Data Platforms","Nat Trust",date(2023,4,10),"Capital Call","Product expansion",30000,0,"Unrealized",None,""],
    ["TX-009","INV-002","Harbor Health Partners","Beaman Ventures",date(2021,9,10),"Capital Call","Initial close",500000,0,"Unrealized",None,""],
    ["TX-010","INV-002","Harbor Health Partners","Kat Trust",date(2021,9,10),"Capital Call","Initial close",250000,0,"Unrealized",None,""],
    ["TX-011","INV-002","Harbor Health Partners","Nat Trust",date(2021,9,10),"Capital Call","Initial close",250000,0,"Unrealized",None,""],
    ["TX-012","INV-002","Harbor Health Partners","Beaman Ventures",date(2022,6,15),"Capital Call","Acquisition tuck-in",250000,0,"Unrealized",None,""],
    ["TX-013","INV-002","Harbor Health Partners","Kat Trust",date(2022,6,15),"Capital Call","Acquisition tuck-in",125000,0,"Unrealized",None,""],
    ["TX-014","INV-002","Harbor Health Partners","Nat Trust",date(2022,6,15),"Capital Call","Acquisition tuck-in",125000,0,"Unrealized",None,""],
    ["TX-015","INV-003","Summit Robotics","Lee Beaman",date(2020,5,5),"Capital Call","Initial close",360000,0,"Unrealized",None,""],
    ["TX-016","INV-003","Summit Robotics","Nat Trust",date(2020,5,5),"Capital Call","Initial close",240000,0,"Unrealized",None,""],
    ["TX-017","INV-003","Summit Robotics","Lee Beaman",date(2021,2,10),"Capital Call","Bridge financing",120000,0,"Unrealized",None,""],
    ["TX-018","INV-003","Summit Robotics","Nat Trust",date(2021,2,10),"Capital Call","Bridge financing",80000,0,"Unrealized",None,""],
    ["TX-019","INV-003","Summit Robotics","Lee Beaman",date(2024,9,30),"Partial Exit","Secondary sale",0,220000,"Realized",date(2024,9,30),""],
    ["TX-020","INV-003","Summit Robotics","Nat Trust",date(2024,9,30),"Partial Exit","Secondary sale",0,150000,"Realized",date(2024,9,30),""],
    ["TX-021","INV-003","Summit Robotics","Lee Beaman",date(2025,2,15),"Dividend","Special dividend",0,30000,"Realized",None,""],
    ["TX-022","INV-003","Summit Robotics","Nat Trust",date(2025,2,15),"Dividend","Special dividend",0,20000,"Realized",None,""],
    ["TX-023","INV-004","Cedar Credit Fund II","Beaman Ventures",date(2022,1,20),"Capital Call","Fund drawdown",150000,0,"Unrealized",None,""],
    ["TX-024","INV-004","Cedar Credit Fund II","Lee Beaman",date(2022,1,20),"Capital Call","Fund drawdown",150000,0,"Unrealized",None,""],
    ["TX-025","INV-004","Cedar Credit Fund II","Kat Trust",date(2022,1,20),"Capital Call","Fund drawdown",150000,0,"Unrealized",None,""],
    ["TX-026","INV-004","Cedar Credit Fund II","Nat Trust",date(2022,1,20),"Capital Call","Fund drawdown",150000,0,"Unrealized",None,""],
    ["TX-027","INV-004","Cedar Credit Fund II","Beaman Ventures",date(2023,1,20),"Capital Call","Follow-on draw",100000,0,"Unrealized",None,""],
    ["TX-028","INV-004","Cedar Credit Fund II","Lee Beaman",date(2023,1,20),"Capital Call","Follow-on draw",100000,0,"Unrealized",None,""],
    ["TX-029","INV-004","Cedar Credit Fund II","Kat Trust",date(2023,1,20),"Capital Call","Follow-on draw",100000,0,"Unrealized",None,""],
    ["TX-030","INV-004","Cedar Credit Fund II","Nat Trust",date(2023,1,20),"Capital Call","Follow-on draw",100000,0,"Unrealized",None,""],
    ["TX-031","INV-004","Cedar Credit Fund II","Beaman Ventures",date(2024,12,15),"Distribution","Quarterly distribution",0,40000,"Realized",None,""],
    ["TX-032","INV-004","Cedar Credit Fund II","Lee Beaman",date(2024,12,15),"Distribution","Quarterly distribution",0,40000,"Realized",None,""],
    ["TX-033","INV-004","Cedar Credit Fund II","Kat Trust",date(2024,12,15),"Distribution","Quarterly distribution",0,40000,"Realized",None,""],
    ["TX-034","INV-004","Cedar Credit Fund II","Nat Trust",date(2024,12,15),"Distribution","Quarterly distribution",0,40000,"Realized",None,""],
    ["TX-035","INV-005","Meridian Consumer Holdings","Beaman Ventures",date(2019,7,1),"Capital Call","Initial close",450000,0,"Unrealized",None,""],
    ["TX-036","INV-005","Meridian Consumer Holdings","Lee Beaman",date(2019,7,1),"Capital Call","Initial close",350000,0,"Unrealized",None,""],
    ["TX-037","INV-005","Meridian Consumer Holdings","Kat Trust",date(2019,7,1),"Capital Call","Initial close",200000,0,"Unrealized",None,""],
    ["TX-038","INV-005","Meridian Consumer Holdings","Beaman Ventures",date(2021,12,31),"Full Exit","Sale to strategic buyer",0,900000,"Realized",date(2021,12,31),""],
    ["TX-039","INV-005","Meridian Consumer Holdings","Lee Beaman",date(2021,12,31),"Full Exit","Sale to strategic buyer",0,700000,"Realized",date(2021,12,31),""],
    ["TX-040","INV-005","Meridian Consumer Holdings","Kat Trust",date(2021,12,31),"Full Exit","Sale to strategic buyer",0,400000,"Realized",date(2021,12,31),""],
    ["TX-041","INV-006","Northshore BioTools","Kat Trust",date(2023,11,3),"Capital Call","Series B",330000,0,"Unrealized",None,""],
    ["TX-042","INV-006","Northshore BioTools","Nat Trust",date(2023,11,3),"Capital Call","Series B",270000,0,"Unrealized",None,""],
    ["TX-043","INV-006","Northshore BioTools","Kat Trust",date(2024,10,15),"Capital Call","Extension round",110000,0,"Unrealized",None,""],
    ["TX-044","INV-006","Northshore BioTools","Nat Trust",date(2024,10,15),"Capital Call","Extension round",90000,0,"Unrealized",None,""],
    ["TX-045","INV-007","Ironclad Energy Services","Beaman Ventures",date(2018,4,12),"Capital Call","Initial close",120000,0,"Unrealized",None,""],
    ["TX-046","INV-007","Ironclad Energy Services","Lee Beaman",date(2018,4,12),"Capital Call","Initial close",80000,0,"Unrealized",None,""],
    ["TX-047","INV-007","Ironclad Energy Services","Kat Trust",date(2018,4,12),"Capital Call","Initial close",100000,0,"Unrealized",None,""],
    ["TX-048","INV-007","Ironclad Energy Services","Nat Trust",date(2018,4,12),"Capital Call","Initial close",100000,0,"Unrealized",None,""],
    ["TX-049","INV-008","Solstice Payments","Beaman Ventures",date(2024,6,18),"Capital Call","Series C",280000,0,"Unrealized",None,""],
    ["TX-050","INV-008","Solstice Payments","Lee Beaman",date(2024,6,18),"Capital Call","Series C",120000,0,"Unrealized",None,""],
    ["TX-051","INV-008","Solstice Payments","Kat Trust",date(2024,6,18),"Capital Call","Series C",200000,0,"Unrealized",None,""],
    ["TX-052","INV-008","Solstice Payments","Nat Trust",date(2024,6,18),"Capital Call","Series C",200000,0,"Unrealized",None,""],
    ["TX-053","INV-008","Solstice Payments","Beaman Ventures",date(2025,5,1),"Capital Call","Growth financing",70000,0,"Unrealized",None,""],
    ["TX-054","INV-008","Solstice Payments","Lee Beaman",date(2025,5,1),"Capital Call","Growth financing",30000,0,"Unrealized",None,""],
    ["TX-055","INV-008","Solstice Payments","Kat Trust",date(2025,5,1),"Capital Call","Growth financing",50000,0,"Unrealized",None,""],
    ["TX-056","INV-008","Solstice Payments","Nat Trust",date(2025,5,1),"Capital Call","Growth financing",50000,0,"Unrealized",None,""],
]

OFFICIAL_MARKS = [
    [date(2024,12,31),"INV-001","Atlas Data Platforms","All Entities",1000000,1400000,"Series C closed above prior round","Priced Round",""],
    [date(2025,12,31),"INV-001","Atlas Data Platforms","All Entities",1400000,1600000,"Secondary transaction and board materials","Priced Round",""],
    [date(2025,12,31),"INV-002","Harbor Health Partners","All Entities",1125000,1125000,"Held at cost pending external marker","Cost",""],
    [date(2025,12,31),"INV-003","Summit Robotics","All Entities",350000,420000,"Secondary transaction reference point","Third-Party Transaction","Partial liquidity achieved"],
    [date(2025,12,31),"INV-004","Cedar Credit Fund II","All Entities",1000000,940000,"Manager statement after distributions","Other","Conservative mark below internal NAV"],
    [date(2021,12,31),"INV-005","Meridian Consumer Holdings","All Entities",1000000,0,"Fully realized exit","Exit Process",""],
    [date(2025,12,31),"INV-006","Northshore BioTools","All Entities",800000,800000,"Held at cost pending next priced round","Cost",""],
    [date(2025,12,31),"INV-007","Ironclad Energy Services","All Entities",400000,0,"Operational impairment","Write-down","Written to zero"],
    [date(2025,12,31),"INV-008","Solstice Payments","All Entities",1000000,1250000,"2025 financing round","Priced Round",""],
]

INTERNAL_MODELS = [
    ["INV-001","Atlas Data Platforms",date(2025,12,31),"Revenue multiple",42000000,0,0,6.2,0,0,0,0,0.08,0.00,None,"Pipeline weighted to enterprise customers"],
    ["INV-002","Harbor Health Partners",date(2025,12,31),"EBITDA multiple",0,18000000,9.5,0,0,0,0,0,0.12,0.00,None,"Margin expansion underwritten internally"],
    ["INV-003","Summit Robotics",date(2025,12,31),"Revenue multiple",21000000,0,0,4.8,0,0,0,0,0.065,0.00,None,"Reflects post-secondary discount"],
    ["INV-004","Cedar Credit Fund II",date(2025,12,31),"NAV / asset-based",0,0,0,0,980000,0,0,0,0.04,0.00,None,"Manager-reported NAV plus accrued income"],
    ["INV-005","Meridian Consumer Holdings",date(2025,12,31),"EBITDA multiple",0,0,0,0,0,0,0,0,0.05,0.00,0,"Realized investment"],
    ["INV-006","Northshore BioTools",date(2025,12,31),"Revenue multiple",24000000,0,0,5.0,0,0,0,0,0.09,0.05,None,"Internal view above official cost"],
    ["INV-007","Ironclad Energy Services",date(2025,12,31),"Manual override",0,0,0,0,0,0,0,0,0.03,0.00,0,"No recovery assumed"],
    ["INV-008","Solstice Payments",date(2025,12,31),"Revenue multiple",36000000,0,0,5.8,0,0,0,0,0.07,0.00,None,"Strong payments growth and improving take rate"],
]

EXIT_SCENARIOS = [
    ["INV-001","Atlas Data Platforms",date(2027,12,31),"Revenue",50000000,7.0,0,0.08,25000000,"Strategic sale scenario"],
    ["INV-002","Harbor Health Partners",date(2027,12,31),"EBITDA",22000000,11.0,0,0.12,18000000,"Sponsor sale after de-leveraging"],
    ["INV-003","Summit Robotics",date(2027,6,30),"EBITDA",14000000,12.0,0,0.065,5000000,"IPO / strategic sale case"],
    ["INV-004","Cedar Credit Fund II",date(2026,12,31),"Other",980000,1.0,0,0.04,0,"Wind-down at current NAV"],
    ["INV-005","Meridian Consumer Holdings",date(2021,12,31),"EBITDA",0,0,0,0.05,0,"Already exited"],
    ["INV-006","Northshore BioTools",date(2028,12,31),"Revenue",32000000,6.0,0,0.09,3000000,"Takeout by strategic diagnostics buyer"],
    ["INV-007","Ironclad Energy Services",date(2025,12,31),"Other",0,0,0,0.03,0,"No upside retained"],
    ["INV-008","Solstice Payments",date(2028,12,31),"Revenue",46000000,7.0,0,0.07,2500000,"M&A exit with modest transaction costs"],
]


def style_cell(cell, *, fill=None, font=None, alignment=None, border=BORDER, number_format=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def style_range_header(ws, row, text, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    style_cell(
        c,
        fill=PatternFill("solid", fgColor=COLOR_DARK_BLUE),
        font=Font(color="FFFFFF", bold=True, size=12),
        alignment=Alignment(horizontal="left"),
    )


def add_table(ws, start_row, end_row, end_col, name):
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.auto_filter.ref = ref


def fit_widths(ws, widths):
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_instructions(ws, title, bullets):
    ws["A1"] = title
    style_cell(
        ws["A1"],
        fill=PatternFill("solid", fgColor=COLOR_DARK_BLUE),
        font=Font(color="FFFFFF", bold=True, size=15),
        alignment=Alignment(horizontal="left"),
    )
    ws.merge_cells("A1:H1")
    ws["A2"] = "Purpose"
    style_cell(
        ws["A2"],
        fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE),
        font=Font(bold=True, color=COLOR_TEXT),
    )
    ws.merge_cells("A2:H2")
    for idx, bullet in enumerate(bullets, start=3):
        ws[f"A{idx}"] = f"• {bullet}"
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)
        ws[f"A{idx}"].alignment = Alignment(wrap_text=True)
    fit_widths(ws, {1: 18})


def create_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.creator = "Codex"
    wb.properties.title = "Private Investment Tracker"
    wb.properties.subject = "Family office private investments and portfolio valuation model"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    cover = wb.create_sheet("Cover & Instructions")
    write_instructions(
        cover,
        "Private Investment & Portfolio Valuation Workbook",
        [
            "Use the yellow cells in the data tabs for manual updates. Formula and output cells are protected visually and should update automatically when the workbook recalculates.",
            "Track all actual capital calls and distributions in the Capital Call & Distribution Ledger. XIRR calculations are based on those dated cash flows plus the selected ending valuation view.",
            "Maintain conservative marks in the Official Valuation Ledger, internal operating marks in the Internal Valuation Model, and upside case values in Exit Scenario Valuation.",
            "Use the Entity Summary Dashboard for entity-level review and the Combined Portfolio Dashboard for overall portfolio reporting.",
            "Refresh formulas in Excel after opening if calculation mode is manual. Filters, tables, and named ranges are already set up for ongoing use.",
        ],
    )
    cover["A10"] = "Entities covered"
    style_cell(cover["A10"], fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    for i, entity in enumerate(ENTITIES, start=11):
        cover[f"A{i}"] = entity
        style_cell(cover[f"A{i}"], fill=PatternFill("solid", fgColor=COLOR_FORMULA))
    cover["D10"] = "Workbook use"
    style_cell(cover["D10"], fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    cover["D11"] = "1. Update Investment Master for new positions and ownership."
    cover["D12"] = "2. Add actual capital calls and distributions in Cash Flow Ledger."
    cover["D13"] = "3. Maintain official, internal, and exit marks on their respective tabs."
    cover["D14"] = "4. Review dashboards and audit checks before distribution."
    for cell_ref in ["D11", "D12", "D13", "D14"]:
        cover[cell_ref].alignment = Alignment(wrap_text=True)
        style_cell(cover[cell_ref], fill=PatternFill("solid", fgColor=COLOR_FORMULA))
    cover["A16"] = "Workbook architecture"
    style_cell(cover["A16"], fill=PatternFill("solid", fgColor=COLOR_DARK_BLUE), font=Font(color="FFFFFF", bold=True))
    cover.merge_cells("A16:H16")
    architecture_headers = ["Tab", "Purpose"]
    for col, value in enumerate(architecture_headers, start=1):
        style_cell(
            cover.cell(row=17, column=col, value=value),
            fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE),
            font=Font(bold=True)
        )
    tab_map = [
        ("Entity Summary Dashboard", "Entity-level snapshot across Beaman Ventures, Lee Beaman, Kat Trust, and Nat Trust."),
        ("Combined Portfolio Dashboard", "Whole-portfolio KPI view and chart package for internal review."),
        ("Investment Master", "One-row-per-investment reference table with ownership and valuation rollups."),
        ("Cash Flow Ledger", "Source-of-truth dated capital calls, distributions, dividends, and exit cash flows."),
        ("Official Valuation Ledger", "Conservative reporting marks anchored to external valuation events."),
        ("Internal Valuation Model", "Team underwriting view using EBITDA, revenue, NAV, or manual override."),
        ("Exit Scenario Valuation", "Upside/downside planning case based on realistic exit assumptions."),
        ("Audit Checks", "Control framework to validate that source data and rollups tie."),
    ]
    for row_idx, (tab_name, purpose) in enumerate(tab_map, start=18):
        cover[f"A{row_idx}"] = tab_name
        cover[f"B{row_idx}"] = purpose
        cover.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        cover[f"B{row_idx}"].alignment = Alignment(wrap_text=True)
        style_cell(cover[f"A{row_idx}"], fill=PatternFill("solid", fgColor=COLOR_FORMULA))
        style_cell(cover[f"B{row_idx}"], fill=PatternFill("solid", fgColor=COLOR_FORMULA))
    fit_widths(cover, {1: 26, 2: 20, 4: 28})
    cover.freeze_panes = "A3"

    assumptions = wb.create_sheet("Assumptions & Methodology")
    write_instructions(
        assumptions,
        "Assumptions & Methodology",
        [
            "Official valuation represents the conservative reporting view and should stay at cost until there is a real external marker event.",
            "Internal valuation reflects the investment team's best estimate using selected valuation methods such as EBITDA, revenue, NAV, or manual override.",
            "Exit scenario value is a planning case and should not be used as official NAV.",
            "Official / Internal / Exit XIRR all use actual dated cash flows from the ledger, plus the relevant ending value as a terminal positive cash flow dated as of today.",
            "TVPI = (Distributions + Residual Value) / Paid-In Capital. DPI = Distributions / Paid-In Capital. RVPI = Residual Value / Paid-In Capital. MOIC is shown as turns.",
            "Entity-level NAV is allocated using current ownership percentages maintained in the Investment Master tab.",
        ],
    )
    assumptions["A10"] = "Valuation views"
    style_cell(assumptions["A10"], fill=PatternFill("solid", fgColor=COLOR_DARK_BLUE), font=Font(color="FFFFFF", bold=True))
    assumptions.merge_cells("A10:H10")
    valuation_blocks = [
        ("Official Conservative Mark", "Board and reporting view. Stay at cost until a priced round, market transaction, or credible external marker supports a change."),
        ("Internal Valuation Mark", "Best internal view of equity value using the selected operating valuation framework and team judgment."),
        ("Potential Exit Value", "Scenario planning view only. Useful for upside framing, capital allocation decisions, and embedded value discussion."),
    ]
    for row_idx, (label, copy) in enumerate(valuation_blocks, start=11):
        assumptions[f"A{row_idx}"] = label
        assumptions[f"B{row_idx}"] = copy
        assumptions.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        style_cell(assumptions[f"A{row_idx}"], fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
        style_cell(assumptions[f"B{row_idx}"], fill=PatternFill("solid", fgColor=COLOR_FORMULA))
        assumptions[f"B{row_idx}"].alignment = Alignment(wrap_text=True)

    lookups = wb.create_sheet("Lookups & Data Validation")
    lookup_sets = {
        "A": ("Entity Names", ENTITIES),
        "B": ("Transaction Types", TRANSACTION_TYPES),
        "C": ("Investment Status", STATUSES),
        "D": ("Valuation Methods", VALUATION_METHODS),
        "E": ("Mark Types", MARK_TYPES),
        "F": ("Sector List", SECTORS),
        "G": ("Asset Type List", ASSET_TYPES),
        "H": ("Yes / No", ["Y", "N"]),
        "I": ("Lead / Co-invest", ["Lead", "Co-invest"]),
    }
    for col_letter, (title, values) in lookup_sets.items():
        lookups[f"{col_letter}1"] = title
        style_cell(lookups[f"{col_letter}1"], fill=PatternFill("solid", fgColor=COLOR_DARK_BLUE), font=Font(color="FFFFFF", bold=True))
        for idx, value in enumerate(values, start=2):
            lookups[f"{col_letter}{idx}"] = value
    fit_widths(lookups, {1: 22, 2: 26, 3: 22, 4: 24, 5: 24, 6: 22, 7: 22, 8: 12, 9: 16})
    lookups.freeze_panes = "A2"

    names = {
        "EntityList": "$A$2:$A$5",
        "TransactionTypeList": f"$B$2:$B${1+len(TRANSACTION_TYPES)}",
        "StatusList": f"$C$2:$C${1+len(STATUSES)}",
        "ValuationMethodList": f"$D$2:$D${1+len(VALUATION_METHODS)}",
        "MarkTypeList": f"$E$2:$E${1+len(MARK_TYPES)}",
        "SectorList": f"$F$2:$F${1+len(SECTORS)}",
        "AssetTypeList": f"$G$2:$G${1+len(ASSET_TYPES)}",
        "YesNoList": "$H$2:$H$3",
        "LeadCoList": "$I$2:$I$3",
    }
    for name, ref in names.items():
        wb.defined_names.add(DefinedName(name, attr_text=f"'Lookups & Data Validation'!{ref}"))

    master = wb.create_sheet("Investment Master")
    master_headers = [
        "Investment ID","Investment Name","Sector / Strategy","Asset Type","Status","Initial Investment Date",
        "Sponsor / Manager","Geography","Official Valuation Method","Internal Valuation Method","Exit Valuation Method",
        "Board Seat?","Lead / Co-invest?","Commitment Amount","Beaman Ventures %","Lee Beaman %","Kat Trust %",
        "Nat Trust %","Ownership % total","Entity ownership notes","Notes","Total invested capital","Total distributions",
        "Official current value","Internal current value","Exit scenario value","Official MOIC","Internal MOIC","Exit MOIC",
        "Official XIRR","Internal XIRR","Exit XIRR"
    ]
    style_range_header(master, 1, "Investment Master | One row per investment", len(master_headers))
    master["A2"] = "Yellow cells are core inputs. Formula columns roll up cash flows, current marks, and performance."
    master.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(master_headers))
    master["A2"].alignment = Alignment(wrap_text=True)
    for col, header in enumerate(master_headers, start=1):
        cell = master.cell(row=3, column=col, value=header)
        style_cell(cell, fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    for idx, inv in enumerate(INVESTMENTS, start=4):
        base = [
            inv["id"], inv["name"], inv["sector"], inv["asset_type"], inv["status"], inv["date"], inv["sponsor"], inv["geo"],
            inv["official_method"], inv["internal_method"], inv["exit_method"], inv["board"], inv["lead"], inv["commitment"],
            inv["bv_pct"], inv["lb_pct"], inv["kt_pct"], inv["nt_pct"], inv["own_total"], inv["entity_notes"], inv["notes"],
        ]
        for col, value in enumerate(base, start=1):
            master.cell(row=idx, column=col, value=value)
        id_ref = f"A{idx}"
        master[f"V{idx}"] = f'=SUMIFS(tblCashFlows[Contribution Amount],tblCashFlows[Investment ID],{id_ref})'
        master[f"W{idx}"] = f'=SUMIFS(tblCashFlows[Distribution Amount],tblCashFlows[Investment ID],{id_ref})'
        master[f"X{idx}"] = f'=XLOOKUP({id_ref},tblOfficialMarks[Investment ID],tblOfficialMarks[New Official Value],0,0,-1)'
        master[f"Y{idx}"] = f'=XLOOKUP({id_ref},tblInternalModel[Investment ID],tblInternalModel[Final Internal Value],0,0,-1)'
        master[f"Z{idx}"] = f'=XLOOKUP({id_ref},tblExitScenarios[Investment ID],tblExitScenarios[Net Exit Value],0,0,-1)'
        master[f"AA{idx}"] = f'=IFERROR((W{idx}+X{idx})/V{idx},"")'
        master[f"AB{idx}"] = f'=IFERROR((W{idx}+Y{idx})/V{idx},"")'
        master[f"AC{idx}"] = f'=IFERROR((W{idx}+Z{idx})/V{idx},"")'
        master[f"AD{idx}"] = f'=IFERROR(LET(cf,FILTER(tblCashFlows[Net Cash Flow],tblCashFlows[Investment ID]={id_ref}),dt,FILTER(tblCashFlows[Transaction Date],tblCashFlows[Investment ID]={id_ref}),XIRR(VSTACK(cf,N(X{idx})),VSTACK(dt,TODAY()))),"")'
        master[f"AE{idx}"] = f'=IFERROR(LET(cf,FILTER(tblCashFlows[Net Cash Flow],tblCashFlows[Investment ID]={id_ref}),dt,FILTER(tblCashFlows[Transaction Date],tblCashFlows[Investment ID]={id_ref}),XIRR(VSTACK(cf,N(Y{idx})),VSTACK(dt,TODAY()))),"")'
        master[f"AF{idx}"] = f'=IFERROR(LET(cf,FILTER(tblCashFlows[Net Cash Flow],tblCashFlows[Investment ID]={id_ref}),dt,FILTER(tblCashFlows[Transaction Date],tblCashFlows[Investment ID]={id_ref}),XIRR(VSTACK(cf,N(Z{idx})),VSTACK(dt,TODAY()))),"")'
    add_table(master, 3, 3 + len(INVESTMENTS), len(master_headers), "tblInvestments")
    master.freeze_panes = "A4"
    fit_widths(master, {1: 14,2: 28,3: 20,4: 16,5: 18,6: 16,7: 24,8: 14,9: 24,10: 24,11: 22,12: 12,13: 14,14: 16,15: 15,16: 13,17: 13,18: 13,19: 15,20: 26,21: 28,22: 18,23: 18,24: 18,25: 18,26: 18,27: 14,28: 14,29: 14,30: 13,31: 13,32: 13})

    ledger = wb.create_sheet("Cash Flow Ledger")
    ledger_headers = [
        "Transaction ID","Investment ID","Investment Name","Entity","Transaction Date","Transaction Type","Description",
        "Contribution Amount","Distribution Amount","Net Cash Flow","Realized / Unrealized flag","Linked valuation date","Notes"
    ]
    style_range_header(ledger, 1, "Capital Call & Distribution Ledger | Enter one dated cash flow per row", len(ledger_headers))
    ledger["A2"] = "Capital calls are negative in Net Cash Flow. Distributions are positive. This ledger feeds all XIRR calculations."
    ledger.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ledger_headers))
    for col, header in enumerate(ledger_headers, start=1):
        style_cell(ledger.cell(row=3, column=col, value=header), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    for idx, row in enumerate(LEDGER, start=4):
        for col, value in enumerate(row, start=1):
            ledger.cell(row=idx, column=col, value=value)
        ledger[f"J{idx}"] = f'=IF(H{idx}="",0,-H{idx})+IF(I{idx}="",0,I{idx})'
    add_table(ledger, 3, 3 + len(LEDGER), len(ledger_headers), "tblCashFlows")
    ledger.freeze_panes = "A4"
    fit_widths(ledger, {1: 14,2: 14,3: 28,4: 18,5: 15,6: 22,7: 24,8: 18,9: 18,10: 15,11: 18,12: 18,13: 24})

    official = wb.create_sheet("Official Valuation Ledger")
    official_headers = [
        "Valuation Date","Investment ID","Investment Name","Entity or All-Entity View","Prior Official Value","New Official Value",
        "Change in Value","Basis for Mark","Mark Type","Notes"
    ]
    style_range_header(official, 1, "Official Valuation Ledger | Conservative reporting marks", len(official_headers))
    official["A2"] = "Default posture is cost until a priced round, transaction, or comparable external marker supports a new official mark."
    official.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(official_headers))
    for col, header in enumerate(official_headers, start=1):
        style_cell(official.cell(row=3, column=col, value=header), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    for idx, row in enumerate(OFFICIAL_MARKS, start=4):
        for col, value in enumerate(row, start=1):
            official.cell(row=idx, column=col, value=value)
        official[f"G{idx}"] = f'=F{idx}-E{idx}'
    add_table(official, 3, 3 + len(OFFICIAL_MARKS), len(official_headers), "tblOfficialMarks")
    official.freeze_panes = "A4"
    fit_widths(official, {1: 15,2: 14,3: 28,4: 20,5: 18,6: 18,7: 16,8: 30,9: 18,10: 24})

    internal = wb.create_sheet("Internal Valuation Model")
    internal_headers = [
        "Investment ID","Investment Name","Valuation Date","Valuation Method","LTM Revenue","LTM EBITDA","EBITDA Multiple","Revenue Multiple",
        "Enterprise Value","Net Debt","Equity Value","Our Ownership %","Our Value","Discount / premium adjustment","Manual override value",
        "Final Internal Value","Official Value","Delta $","Delta %","Notes"
    ]
    style_range_header(internal, 1, "Internal Valuation Model | Team view of intrinsic value", len(internal_headers))
    internal["A2"] = "Use the selected method to drive value. Manual override takes precedence only when that method is selected."
    internal.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(internal_headers))
    for col, header in enumerate(internal_headers, start=1):
        style_cell(internal.cell(row=3, column=col, value=header), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    for idx, row in enumerate(INTERNAL_MODELS, start=4):
        for col, value in enumerate(row, start=1):
            internal.cell(row=idx, column=col, value=value)
        internal[f"I{idx}"] = f'=IF(D{idx}="EBITDA multiple",F{idx}*G{idx},IF(D{idx}="Revenue multiple",E{idx}*H{idx},IF(D{idx}="NAV / asset-based",E{idx},IF(D{idx}="Manual override",O{idx},0))))'
        internal[f"K{idx}"] = f'=IF(D{idx}="NAV / asset-based",E{idx},I{idx}-J{idx})'
        internal[f"M{idx}"] = f'=K{idx}*L{idx}'
        internal[f"P{idx}"] = f'=IF(D{idx}="Manual override",O{idx},M{idx}*(1+N{idx}))'
        internal[f"Q{idx}"] = f'=XLOOKUP(A{idx},tblOfficialMarks[Investment ID],tblOfficialMarks[New Official Value],0,0,-1)'
        internal[f"R{idx}"] = f'=P{idx}-Q{idx}'
        internal[f"S{idx}"] = f'=IFERROR(R{idx}/Q{idx},"")'
    add_table(internal, 3, 3 + len(INTERNAL_MODELS), len(internal_headers), "tblInternalModel")
    internal.freeze_panes = "A4"
    fit_widths(internal, {1: 14,2: 28,3: 15,4: 20,5: 16,6: 16,7: 16,8: 16,9: 18,10: 12,11: 16,12: 14,13: 16,14: 18,15: 18,16: 18,17: 16,18: 14,19: 12,20: 26})

    exit_ws = wb.create_sheet("Exit Scenario Valuation")
    exit_headers = [
        "Investment ID","Investment Name","Exit Date Assumption","Exit Metric","Exit Metric Amount","Exit Multiple",
        "Enterprise Value","Less Debt","Equity Value","Our Ownership %","Gross Exit Value","Estimated Fees / Taxes / Costs",
        "Net Exit Value","Official Value","Internal Value","Upside vs Official","Upside vs Internal","Notes"
    ]
    style_range_header(exit_ws, 1, "Exit Scenario Valuation | Planning case only", len(exit_headers))
    exit_ws["A2"] = "Exit values support scenario planning and should not be used as official NAV."
    exit_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(exit_headers))
    for col, header in enumerate(exit_headers, start=1):
        style_cell(exit_ws.cell(row=3, column=col, value=header), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    for idx, row in enumerate(EXIT_SCENARIOS, start=4):
        for col, value in enumerate(row, start=1):
            exit_ws.cell(row=idx, column=col, value=value)
        exit_ws[f"G{idx}"] = f'=IF(D{idx}="Other",E{idx},E{idx}*F{idx})'
        exit_ws[f"I{idx}"] = f'=G{idx}-H{idx}'
        exit_ws[f"K{idx}"] = f'=I{idx}*J{idx}'
        exit_ws[f"M{idx}"] = f'=K{idx}-L{idx}'
        exit_ws[f"N{idx}"] = f'=XLOOKUP(A{idx},tblOfficialMarks[Investment ID],tblOfficialMarks[New Official Value],0,0,-1)'
        exit_ws[f"O{idx}"] = f'=XLOOKUP(A{idx},tblInternalModel[Investment ID],tblInternalModel[Final Internal Value],0,0,-1)'
        exit_ws[f"P{idx}"] = f'=M{idx}-N{idx}'
        exit_ws[f"Q{idx}"] = f'=M{idx}-O{idx}'
    add_table(exit_ws, 3, 3 + len(EXIT_SCENARIOS), len(exit_headers), "tblExitScenarios")
    exit_ws.freeze_panes = "A4"
    fit_widths(exit_ws, {1: 14,2: 28,3: 16,4: 14,5: 18,6: 14,7: 18,8: 14,9: 16,10: 14,11: 18,12: 22,13: 18,14: 16,15: 16,16: 16,17: 16,18: 24})

    entity_perf = wb.create_sheet("Entity-Level Performance")
    entity_headers = [
        "Entity","Total invested capital","Total distributions","Remaining cost basis","Official NAV","Internal NAV",
        "Exit Scenario NAV","TVPI","DPI","RVPI","Official XIRR","Internal XIRR","Exit Scenario XIRR",
        "Number of active investments","Number of realized investments"
    ]
    style_range_header(entity_perf, 1, "Entity-Level Performance | Snapshot", len(entity_headers))
    for col, header in enumerate(entity_headers, start=1):
        style_cell(entity_perf.cell(row=3, column=col, value=header), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    ownership_choice = 'CHOOSE(MATCH(A{row},EntityList,0),tblInvestments[Beaman Ventures %],tblInvestments[Lee Beaman %],tblInvestments[Kat Trust %],tblInvestments[Nat Trust %])'
    for idx, entity in enumerate(ENTITIES, start=4):
        entity_perf[f"A{idx}"] = entity
        entity_perf[f"B{idx}"] = f'=SUMIFS(tblCashFlows[Contribution Amount],tblCashFlows[Entity],A{idx})'
        entity_perf[f"C{idx}"] = f'=SUMIFS(tblCashFlows[Distribution Amount],tblCashFlows[Entity],A{idx})'
        entity_perf[f"D{idx}"] = f'=MAX(0,B{idx}-C{idx})'
        choice = ownership_choice.format(row=idx)
        entity_perf[f"E{idx}"] = f'=SUMPRODUCT({choice},tblInvestments[Official current value])'
        entity_perf[f"F{idx}"] = f'=SUMPRODUCT({choice},tblInvestments[Internal current value])'
        entity_perf[f"G{idx}"] = f'=SUMPRODUCT({choice},tblInvestments[Exit scenario value])'
        entity_perf[f"H{idx}"] = f'=IFERROR((C{idx}+E{idx})/B{idx},"")'
        entity_perf[f"I{idx}"] = f'=IFERROR(C{idx}/B{idx},"")'
        entity_perf[f"J{idx}"] = f'=IFERROR(E{idx}/B{idx},"")'
        entity_perf[f"K{idx}"] = f'=IFERROR(LET(cf,FILTER(tblCashFlows[Net Cash Flow],tblCashFlows[Entity]=A{idx}),dt,FILTER(tblCashFlows[Transaction Date],tblCashFlows[Entity]=A{idx}),XIRR(VSTACK(cf,N(E{idx})),VSTACK(dt,TODAY()))),"")'
        entity_perf[f"L{idx}"] = f'=IFERROR(LET(cf,FILTER(tblCashFlows[Net Cash Flow],tblCashFlows[Entity]=A{idx}),dt,FILTER(tblCashFlows[Transaction Date],tblCashFlows[Entity]=A{idx}),XIRR(VSTACK(cf,N(F{idx})),VSTACK(dt,TODAY()))),"")'
        entity_perf[f"M{idx}"] = f'=IFERROR(LET(cf,FILTER(tblCashFlows[Net Cash Flow],tblCashFlows[Entity]=A{idx}),dt,FILTER(tblCashFlows[Transaction Date],tblCashFlows[Entity]=A{idx}),XIRR(VSTACK(cf,N(G{idx})),VSTACK(dt,TODAY()))),"")'
        entity_perf[f"N{idx}"] = f'=SUMPRODUCT(--({choice}>0),--((tblInvestments[Status]="Active")+(tblInvestments[Status]="Partially Realized")>0))'
        entity_perf[f"O{idx}"] = f'=SUMPRODUCT(--({choice}>0),--(tblInvestments[Status]="Realized"))'
    add_table(entity_perf, 3, 7, len(entity_headers), "tblEntityPerf")
    entity_perf["A10"] = "Annualized view by calendar year"
    style_cell(entity_perf["A10"], fill=PatternFill("solid", fgColor=COLOR_DARK_BLUE), font=Font(color="FFFFFF", bold=True))
    annual_headers = ["Year","Entity","Contributions","Distributions","Net Cash Flow"]
    for col, header in enumerate(annual_headers, start=1):
        style_cell(entity_perf.cell(row=11, column=col, value=header), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    years = list(range(2018, 2027))
    r = 12
    for year in years:
        for entity in ENTITIES:
            entity_perf[f"A{r}"] = year
            entity_perf[f"B{r}"] = entity
            entity_perf[f"C{r}"] = f'=SUMIFS(tblCashFlows[Contribution Amount],tblCashFlows[Entity],B{r},tblCashFlows[Transaction Date],">="&DATE(A{r},1,1),tblCashFlows[Transaction Date],"<"&DATE(A{r}+1,1,1))'
            entity_perf[f"D{r}"] = f'=SUMIFS(tblCashFlows[Distribution Amount],tblCashFlows[Entity],B{r},tblCashFlows[Transaction Date],">="&DATE(A{r},1,1),tblCashFlows[Transaction Date],"<"&DATE(A{r}+1,1,1))'
            entity_perf[f"E{r}"] = f'=D{r}-C{r}'
            r += 1
    add_table(entity_perf, 11, r - 1, len(annual_headers), "tblEntityAnnual")
    entity_perf.freeze_panes = "A4"
    fit_widths(entity_perf, {1: 18,2: 18,3: 18,4: 18,5: 16,6: 16,7: 18,8: 12,9: 12,10: 12,11: 14,12: 14,13: 16,14: 18,15: 18})

    inv_perf = wb.create_sheet("Investment-Level Performance")
    inv_headers = [
        "Investment ID","Investment Name","Status","Total Invested Capital","Total Distributions","Official Value","Internal Value","Exit Value",
        "Official TVPI","Internal TVPI","Exit TVPI","Official XIRR","Internal XIRR","Exit XIRR","Unrealized Gain / Loss","Realized Gain / Loss",
        "Total Gain / Loss","Holding Period","Which entities own it","Stale mark?"
    ]
    style_range_header(inv_perf, 1, "Investment-Level Performance | One row per investment", len(inv_headers))
    for col, header in enumerate(inv_headers, start=1):
        style_cell(inv_perf.cell(row=3, column=col, value=header), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    for idx, inv in enumerate(INVESTMENTS, start=4):
        inv_perf[f"A{idx}"] = inv["id"]
        inv_perf[f"B{idx}"] = f'=XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Investment Name],"")'
        inv_perf[f"C{idx}"] = f'=XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Status],"")'
        inv_perf[f"D{idx}"] = f'=XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Total invested capital],0)'
        inv_perf[f"E{idx}"] = f'=XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Total distributions],0)'
        inv_perf[f"F{idx}"] = f'=XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Official current value],0)'
        inv_perf[f"G{idx}"] = f'=XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Internal current value],0)'
        inv_perf[f"H{idx}"] = f'=XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Exit scenario value],0)'
        inv_perf[f"I{idx}"] = f'=IFERROR((E{idx}+F{idx})/D{idx},"")'
        inv_perf[f"J{idx}"] = f'=IFERROR((E{idx}+G{idx})/D{idx},"")'
        inv_perf[f"K{idx}"] = f'=IFERROR((E{idx}+H{idx})/D{idx},"")'
        inv_perf[f"L{idx}"] = f'=XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Official XIRR],"")'
        inv_perf[f"M{idx}"] = f'=XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Internal XIRR],"")'
        inv_perf[f"N{idx}"] = f'=XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Exit XIRR],"")'
        inv_perf[f"O{idx}"] = f'=F{idx}-MAX(0,D{idx}-E{idx})'
        inv_perf[f"P{idx}"] = f'=MAX(0,E{idx}-D{idx})'
        inv_perf[f"Q{idx}"] = f'=E{idx}+F{idx}-D{idx}'
        inv_perf[f"R{idx}"] = f'=YEARFRAC(XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Initial Investment Date],TODAY()),TODAY())'
        inv_perf[f"S{idx}"] = f'=TEXTJOIN(", ",TRUE,IF(XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Beaman Ventures %],0)>0,"Beaman Ventures",""),IF(XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Lee Beaman %],0)>0,"Lee Beaman",""),IF(XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Kat Trust %],0)>0,"Kat Trust",""),IF(XLOOKUP(A{idx},tblInvestments[Investment ID],tblInvestments[Nat Trust %],0)>0,"Nat Trust",""))'
        inv_perf[f"T{idx}"] = f'=IF(TODAY()-XLOOKUP(A{idx},tblOfficialMarks[Investment ID],tblOfficialMarks[Valuation Date],TODAY(),0,-1)>120,"Update mark","Current")'
    add_table(inv_perf, 3, 3 + len(INVESTMENTS), len(inv_headers), "tblInvestmentPerf")
    inv_perf.freeze_panes = "A4"
    fit_widths(inv_perf, {1: 14,2: 28,3: 18,4: 18,5: 18,6: 16,7: 16,8: 14,9: 12,10: 12,11: 12,12: 12,13: 12,14: 12,15: 18,16: 16,17: 16,18: 14,19: 28,20: 14})

    cash_timeline = wb.create_sheet("Cash Flow Timeline")
    timeline_headers = ["Month"]
    for entity in ENTITIES:
        timeline_headers.extend([f"{entity} Contributions", f"{entity} Distributions", f"{entity} Net"])
    timeline_headers.extend(["Total Contributions", "Total Distributions", "Total Net"])
    style_range_header(cash_timeline, 1, "Cash Flow Timeline | Monthly liquidity view", len(timeline_headers))
    for col, header in enumerate(timeline_headers, start=1):
        style_cell(cash_timeline.cell(row=3, column=col, value=header), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    timeline_months = []
    start = date(2018, 4, 1)
    end = date(2026, 12, 1)
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        timeline_months.append(date(y, m, 1))
        if m == 12:
            y += 1
            m = 1
        else:
            m += 1
    row = 4
    for month in timeline_months:
        cash_timeline.cell(row=row, column=1, value=month)
        col = 2
        for entity in ENTITIES:
            cash_timeline.cell(row=row, column=col, value=f'=SUMIFS(tblCashFlows[Contribution Amount],tblCashFlows[Entity],"{entity}",tblCashFlows[Transaction Date],">="&A{row},tblCashFlows[Transaction Date],"<"&EDATE(A{row},1))')
            cash_timeline.cell(row=row, column=col + 1, value=f'=SUMIFS(tblCashFlows[Distribution Amount],tblCashFlows[Entity],"{entity}",tblCashFlows[Transaction Date],">="&A{row},tblCashFlows[Transaction Date],"<"&EDATE(A{row},1))')
            cash_timeline.cell(row=row, column=col + 2, value=f'={get_column_letter(col+1)}{row}-{get_column_letter(col)}{row}')
            col += 3
        cash_timeline.cell(row=row, column=col, value=f'=SUM(B{row},E{row},H{row},K{row})')
        cash_timeline.cell(row=row, column=col + 1, value=f'=SUM(C{row},F{row},I{row},L{row})')
        cash_timeline.cell(row=row, column=col + 2, value=f'={get_column_letter(col+1)}{row}-{get_column_letter(col)}{row}')
        row += 1
    add_table(cash_timeline, 3, row - 1, len(timeline_headers), "tblCashTimeline")
    cash_timeline.freeze_panes = "A4"
    fit_widths(cash_timeline, {1: 14, 2: 18, 3: 18, 4: 14, 5: 18, 6: 18, 7: 14, 8: 18, 9: 18, 10: 14, 11: 18, 12: 18, 13: 14, 14: 18, 15: 18, 16: 14})

    entity_dash = wb.create_sheet("Entity Summary Dashboard")
    entity_dash["A1"] = "Entity Summary Dashboard"
    style_cell(entity_dash["A1"], fill=PatternFill("solid", fgColor=COLOR_DARK_BLUE), font=Font(color="FFFFFF", bold=True, size=15))
    entity_dash.merge_cells("A1:H1")
    kpis = [
        ("Total Invested Capital", "=SUM(tblEntityPerf[Total invested capital])"),
        ("Total Distributions", "=SUM(tblEntityPerf[Total distributions])"),
        ("Official NAV", "=SUM(tblEntityPerf[Official NAV])"),
        ("Internal NAV", "=SUM(tblEntityPerf[Internal NAV])"),
        ("Exit Scenario Value", "=SUM(tblEntityPerf[Exit Scenario NAV])"),
        ("Official XIRR", '=IFERROR(LET(cf,tblCashFlows[Net Cash Flow],dt,tblCashFlows[Transaction Date],tv,SUM(tblEntityPerf[Official NAV]),XIRR(VSTACK(cf,tv),VSTACK(dt,TODAY()))),"")'),
    ]
    for i, (label, formula) in enumerate(kpis, start=3):
        entity_dash[f"A{i}"] = label
        entity_dash[f"B{i}"] = formula
        style_cell(entity_dash[f"A{i}"], fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
        style_cell(entity_dash[f"B{i}"], fill=PatternFill("solid", fgColor=COLOR_GREEN), font=Font(bold=True))
    entity_dash["D3"] = "By entity snapshot"
    style_cell(entity_dash["D3"], fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    for r in range(4, 8):
        for c, source_col in enumerate(["A","B","C","E","F","G","H","K"], start=4):
            entity_dash.cell(row=r, column=c, value=f"='Entity-Level Performance'!{source_col}{r}")
    headers = ["Entity","Invested","Distributed","Official NAV","Internal NAV","Exit NAV","TVPI","Official XIRR"]
    for c, h in enumerate(headers, start=4):
        style_cell(entity_dash.cell(row=4, column=c, value=h), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    for src, dest in zip(range(4, 8), range(5, 9)):
        entity_dash[f"D{dest}"] = f"='Entity-Level Performance'!A{src}"
        entity_dash[f"E{dest}"] = f"='Entity-Level Performance'!B{src}"
        entity_dash[f"F{dest}"] = f"='Entity-Level Performance'!C{src}"
        entity_dash[f"G{dest}"] = f"='Entity-Level Performance'!E{src}"
        entity_dash[f"H{dest}"] = f"='Entity-Level Performance'!F{src}"
        entity_dash[f"I{dest}"] = f"='Entity-Level Performance'!G{src}"
        entity_dash[f"J{dest}"] = f"='Entity-Level Performance'!H{src}"
        entity_dash[f"K{dest}"] = f"='Entity-Level Performance'!K{src}"
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Official vs Internal vs Exit NAV by Entity"
    chart.y_axis.title = "Entity"
    chart.x_axis.title = "Value"
    data = Reference(entity_dash, min_col=7, max_col=9, min_row=4, max_row=8)
    cats = Reference(entity_dash, min_col=4, min_row=5, max_row=8)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 14
    entity_dash.add_chart(chart, "D11")
    fit_widths(entity_dash, {1: 24,2: 18,4: 18,5: 16,6: 16,7: 16,8: 16,9: 16,10: 12,11: 12})

    combined_dash = wb.create_sheet("Combined Portfolio Dashboard")
    combined_dash["A1"] = "Combined Portfolio Dashboard"
    style_cell(combined_dash["A1"], fill=PatternFill("solid", fgColor=COLOR_DARK_BLUE), font=Font(color="FFFFFF", bold=True, size=15))
    combined_dash.merge_cells("A1:H1")
    portfolio_kpis = [
        ("Total Invested Capital", "=SUM(tblInvestments[Total invested capital])"),
        ("Total Distributions", "=SUM(tblInvestments[Total distributions])"),
        ("Official NAV", "=SUM(tblInvestments[Official current value])"),
        ("Internal NAV", "=SUM(tblInvestments[Internal current value])"),
        ("Exit Scenario Value", "=SUM(tblInvestments[Exit scenario value])"),
        ("Official TVPI", '=IFERROR((SUM(tblInvestments[Total distributions])+SUM(tblInvestments[Official current value]))/SUM(tblInvestments[Total invested capital]),"")'),
        ("Internal TVPI", '=IFERROR((SUM(tblInvestments[Total distributions])+SUM(tblInvestments[Internal current value]))/SUM(tblInvestments[Total invested capital]),"")'),
        ("Exit TVPI", '=IFERROR((SUM(tblInvestments[Total distributions])+SUM(tblInvestments[Exit scenario value]))/SUM(tblInvestments[Total invested capital]),"")'),
        ("Official XIRR", '=IFERROR(LET(cf,tblCashFlows[Net Cash Flow],dt,tblCashFlows[Transaction Date],tv,SUM(tblInvestments[Official current value]),XIRR(VSTACK(cf,tv),VSTACK(dt,TODAY()))),"")'),
        ("Internal XIRR", '=IFERROR(LET(cf,tblCashFlows[Net Cash Flow],dt,tblCashFlows[Transaction Date],tv,SUM(tblInvestments[Internal current value]),XIRR(VSTACK(cf,tv),VSTACK(dt,TODAY()))),"")'),
        ("Exit XIRR", '=IFERROR(LET(cf,tblCashFlows[Net Cash Flow],dt,tblCashFlows[Transaction Date],tv,SUM(tblInvestments[Exit scenario value]),XIRR(VSTACK(cf,tv),VSTACK(dt,TODAY()))),"")'),
        ("Realized Value", "=SUM(tblInvestments[Total distributions])"),
        ("Unrealized Value", "=SUM(tblInvestments[Official current value])"),
        ("Embedded Value Creation", "=SUM(tblInvestments[Internal current value])-SUM(tblInvestments[Official current value])"),
        ("Exit Upside", "=SUM(tblInvestments[Exit scenario value])-SUM(tblInvestments[Internal current value])"),
        ("Number of Investments", "=ROWS(tblInvestments[Investment ID])"),
        ("Number of Active Investments", '=COUNTIF(tblInvestments[Status],"Active")+COUNTIF(tblInvestments[Status],"Partially Realized")'),
        ("Number of Realized Investments", '=COUNTIF(tblInvestments[Status],"Realized")'),
    ]
    for idx, (label, formula) in enumerate(portfolio_kpis, start=3):
        col = 1 if idx <= 11 else 4
        row = idx if idx <= 11 else idx - 9
        combined_dash.cell(row=row, column=col, value=label)
        combined_dash.cell(row=row, column=col + 1, value=formula)
        style_cell(combined_dash.cell(row=row, column=col), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
        style_cell(combined_dash.cell(row=row, column=col + 1), fill=PatternFill("solid", fgColor=COLOR_GREEN), font=Font(bold=True))
    combined_dash["F3"] = "Top investments by value"
    style_cell(combined_dash["F3"], fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    top_headers = ["Investment","Official Value","Internal Value","Exit Value","Gain / Loss"]
    for c, h in enumerate(top_headers, start=6):
        style_cell(combined_dash.cell(row=4, column=c, value=h), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    for idx in range(4, 12):
        src = idx
        combined_dash[f"F{idx+1}"] = f"='Investment-Level Performance'!B{src}"
        combined_dash[f"G{idx+1}"] = f"='Investment-Level Performance'!F{src}"
        combined_dash[f"H{idx+1}"] = f"='Investment-Level Performance'!G{src}"
        combined_dash[f"I{idx+1}"] = f"='Investment-Level Performance'!H{src}"
        combined_dash[f"J{idx+1}"] = f"='Investment-Level Performance'!Q{src}"
    chart2 = BarChart()
    chart2.title = "Top Investments by Official / Internal / Exit Value"
    chart2.style = 10
    chart2.y_axis.title = "Value"
    data2 = Reference(combined_dash, min_col=7, max_col=9, min_row=4, max_row=12)
    cats2 = Reference(combined_dash, min_col=6, min_row=5, max_row=12)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.height = 8
    chart2.width = 15
    combined_dash.add_chart(chart2, "F15")
    line = LineChart()
    line.title = "Capital Called Over Time"
    line.style = 2
    line.y_axis.title = "Net Cash Flow"
    data3 = Reference(cash_timeline, min_col=14, max_col=16, min_row=3, max_row=min(39, 3 + len(timeline_months)))
    cats3 = Reference(cash_timeline, min_col=1, min_row=4, max_row=min(39, 3 + len(timeline_months)))
    line.add_data(data3, titles_from_data=True)
    line.set_categories(cats3)
    line.height = 8
    line.width = 15
    combined_dash.add_chart(line, "A16")
    fit_widths(combined_dash, {1: 24,2: 18,4: 24,5: 18,6: 24,7: 16,8: 16,9: 16,10: 14})

    audit = wb.create_sheet("Audit Checks")
    audit_headers = ["Control","Result","Status"]
    style_range_header(audit, 1, "Audit Checks", len(audit_headers))
    for c, h in enumerate(audit_headers, start=1):
        style_cell(audit.cell(row=3, column=c, value=h), fill=PatternFill("solid", fgColor=COLOR_LIGHT_BLUE), font=Font(bold=True))
    checks = [
        ("Invested capital ledger ties to investment totals", "=SUM(tblCashFlows[Contribution Amount])-SUM(tblInvestments[Total invested capital])", '=IF(ABS(B4)<0.01,"PASS","FAIL")'),
        ("Distributions ledger ties to investment totals", "=SUM(tblCashFlows[Distribution Amount])-SUM(tblInvestments[Total distributions])", '=IF(ABS(B5)<0.01,"PASS","FAIL")'),
        ("Official marks tie to investment totals", "=SUM(tblOfficialMarks[New Official Value])-SUM(tblInvestments[Official current value])", '=IF(ABS(B6)<0.01,"PASS","FAIL")'),
        ("Internal values tie to dashboard totals", "=SUM(tblInternalModel[Final Internal Value])-SUM(tblInvestments[Internal current value])", '=IF(ABS(B7)<0.01,"PASS","FAIL")'),
        ("Exit values tie to dashboard totals", "=SUM(tblExitScenarios[Net Exit Value])-SUM(tblInvestments[Exit scenario value])", '=IF(ABS(B8)<0.01,"PASS","FAIL")'),
        ("Entity totals tie to combined official total", "=SUM(tblEntityPerf[Official NAV])-SUM(tblInvestments[Official current value])", '=IF(ABS(B9)<0.01,"PASS","FAIL")'),
        ("No missing investment IDs", '=COUNTIF(tblCashFlows[Investment ID],"")+COUNTIF(tblOfficialMarks[Investment ID],"")+COUNTIF(tblInternalModel[Investment ID],"")+COUNTIF(tblExitScenarios[Investment ID],"")', '=IF(B10=0,"PASS","FAIL")'),
        ("No invalid entity names in ledger", '=SUMPRODUCT(--ISNA(MATCH(tblCashFlows[Entity],EntityList,0)))', '=IF(B11=0,"PASS","FAIL")'),
        ("No blank transaction dates in ledger", '=COUNTBLANK(tblCashFlows[Transaction Date])', '=IF(B12=0,"PASS","FAIL")'),
        ("No duplicate transaction IDs", '=ROWS(tblCashFlows[Transaction ID])-COUNTA(UNIQUE(tblCashFlows[Transaction ID]))', '=IF(B13=0,"PASS","FAIL")'),
        ("No missing valuation methods on master", '=COUNTBLANK(tblInvestments[Internal Valuation Method])+COUNTBLANK(tblInvestments[Official Valuation Method])+COUNTBLANK(tblInvestments[Exit Valuation Method])', '=IF(B14=0,"PASS","FAIL")'),
        ("Investments with cash flows but no official mark", '=SUMPRODUCT(--(COUNTIF(tblCashFlows[Investment ID],tblInvestments[Investment ID])>0),--ISNA(MATCH(tblInvestments[Investment ID],tblOfficialMarks[Investment ID],0)))', '=IF(B15=0,"PASS","FAIL")'),
    ]
    for idx, (label, result, status) in enumerate(checks, start=4):
        audit[f"A{idx}"] = label
        audit[f"B{idx}"] = result
        audit[f"C{idx}"] = status
    add_table(audit, 3, 3 + len(checks), 3, "tblAuditChecks")
    audit.freeze_panes = "A4"
    fit_widths(audit, {1: 44, 2: 18, 3: 12})

    # Formats and data validation
    percent_cols = {
        "Investment Master": ["O","P","Q","R","S","AD","AE","AF"],
        "Internal Valuation Model": ["L","N","S"],
        "Exit Scenario Valuation": ["J"],
        "Entity-Level Performance": ["H","I","J","K","L","M"],
        "Investment-Level Performance": ["I","J","K","L","M","N"],
    }
    money_sheets = {
        "Investment Master": list("NVWXYZ"),
        "Cash Flow Ledger": ["H","I","J"],
        "Official Valuation Ledger": ["E","F","G"],
        "Internal Valuation Model": ["E","F","G","H","I","J","K","M","O","P","Q","R"],
        "Exit Scenario Valuation": ["E","F","G","H","I","K","L","M","N","O","P","Q"],
        "Entity-Level Performance": ["B","C","D","E","F","G"],
        "Investment-Level Performance": ["D","E","F","G","H","O","P","Q"],
        "Cash Flow Timeline": [get_column_letter(i) for i in range(2, 17)],
    }
    for sheet_name, cols in money_sheets.items():
        ws = wb[sheet_name]
        for col in cols:
            for row in range(4, ws.max_row + 1):
                ws[f"{col}{row}"].number_format = '$#,##0'
    for sheet_name, cols in percent_cols.items():
        ws = wb[sheet_name]
        for col in cols:
            for row in range(4, ws.max_row + 1):
                ws[f"{col}{row}"].number_format = '0.0%'

    for ws in [master, ledger, official, internal, exit_ws]:
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                if cell.column <= 21 if ws.title == "Investment Master" else (
                    True if ws.title == "Capital Call & Distribution Ledger" and cell.column in [1,2,3,4,5,6,7,11,12,13] else False
                ):
                    pass
        # selected columns styled below

    yellow_fill = PatternFill("solid", fgColor=COLOR_YELLOW)
    gray_fill = PatternFill("solid", fgColor=COLOR_FORMULA)
    green_fill = PatternFill("solid", fgColor=COLOR_GREEN)

    for row in range(4, master.max_row + 1):
        for col in range(1, 22):
            style_cell(master.cell(row=row, column=col), fill=yellow_fill if col <= 21 else gray_fill)
        for col in range(22, 33):
            style_cell(master.cell(row=row, column=col), fill=gray_fill)

    for row in range(4, ledger.max_row + 1):
        for col in range(1, 10):
            style_cell(ledger.cell(row=row, column=col), fill=yellow_fill if col != 10 else gray_fill)
        for col in range(10, 14):
            style_cell(ledger.cell(row=row, column=col), fill=gray_fill if col == 10 else yellow_fill)

    for ws in [official, internal, exit_ws]:
        for row in range(4, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                fill = gray_fill
                if ws.title == "Official Valuation Ledger" and col in [1,2,3,4,5,6,8,9,10]:
                    fill = yellow_fill
                if ws.title == "Internal Valuation Model" and col in [1,2,3,4,5,6,7,8,10,12,14,15,20]:
                    fill = yellow_fill
                if ws.title == "Exit Scenario Valuation" and col in [1,2,3,4,5,6,8,10,12,18]:
                    fill = yellow_fill
                style_cell(ws.cell(row=row, column=col), fill=fill)

    for ws in [entity_perf, inv_perf, cash_timeline, audit]:
        for row in range(4, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                style_cell(ws.cell(row=row, column=col), fill=gray_fill)

    # Make dashboard output cells green
    for ws_name in ["Entity Summary Dashboard", "Combined Portfolio Dashboard"]:
        ws = wb[ws_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.column in (2, 5) and cell.row >= 3 and cell.value is not None:
                    style_cell(cell, fill=green_fill, font=Font(bold=True))

    # Data validation
    dv_entity = DataValidation(type="list", formula1="=EntityList", allow_blank=False)
    dv_status = DataValidation(type="list", formula1="=StatusList", allow_blank=False)
    dv_sector = DataValidation(type="list", formula1="=SectorList", allow_blank=True)
    dv_asset = DataValidation(type="list", formula1="=AssetTypeList", allow_blank=True)
    dv_master_methods = DataValidation(type="list", formula1="=ValuationMethodList", allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1="=YesNoList", allow_blank=True)
    dv_lead = DataValidation(type="list", formula1="=LeadCoList", allow_blank=True)
    for dv in [dv_entity, dv_status, dv_sector, dv_asset, dv_master_methods, dv_yesno, dv_lead]:
        master.add_data_validation(dv)
    dv_sector.add(f"C4:C{master.max_row}")
    dv_asset.add(f"D4:D{master.max_row}")
    dv_status.add(f"E4:E{master.max_row}")
    dv_master_methods.add(f"I4:K{master.max_row}")
    dv_yesno.add(f"L4:L{master.max_row}")
    dv_lead.add(f"M4:M{master.max_row}")

    dv_entity_form = DataValidation(type="list", formula1="=EntityList", allow_blank=False)
    master.add_data_validation(dv_entity_form)

    dv_ledger_entity = DataValidation(type="list", formula1="=EntityList", allow_blank=False)
    dv_txn = DataValidation(type="list", formula1="=TransactionTypeList", allow_blank=False)
    ledger.add_data_validation(dv_ledger_entity)
    ledger.add_data_validation(dv_txn)
    dv_ledger_entity.add(f"D4:D{ledger.max_row}")
    dv_txn.add(f"F4:F{ledger.max_row}")

    dv_mark = DataValidation(type="list", formula1="=MarkTypeList", allow_blank=False)
    official.add_data_validation(dv_mark)
    dv_mark.add(f"I4:I{official.max_row}")
    dv_internal_methods = DataValidation(type="list", formula1="=ValuationMethodList", allow_blank=True)
    internal.add_data_validation(dv_internal_methods)
    dv_internal_methods.add(f"D4:D{internal.max_row}")
    exit_method_dv = DataValidation(type="list", formula1='"Revenue,EBITDA,Other"', allow_blank=True)
    exit_ws.add_data_validation(exit_method_dv)
    exit_method_dv.add(f"D4:D{exit_ws.max_row}")

    # Conditional formatting
    inv_perf.conditional_formatting.add(
        f"I4:I{inv_perf.max_row}",
        CellIsRule(operator="greaterThan", formula=["2"], fill=green_fill),
    )
    inv_perf.conditional_formatting.add(
        f"F4:F{inv_perf.max_row}",
        FormulaRule(formula=["F4<D4-E4"], fill=PatternFill("solid", fgColor="F4CCCC")),
    )
    inv_perf.conditional_formatting.add(
        f"C4:C{inv_perf.max_row}",
        FormulaRule(formula=['$C4="Realized"'], fill=PatternFill("solid", fgColor="D9EAD3")),
    )
    inv_perf.conditional_formatting.add(
        f"T4:T{inv_perf.max_row}",
        FormulaRule(formula=['$T4="Update mark"'], fill=PatternFill("solid", fgColor="FCE5CD")),
    )
    audit.conditional_formatting.add(
        f"C4:C{audit.max_row}",
        FormulaRule(formula=['$C4="PASS"'], fill=PatternFill("solid", fgColor="D9EAD3")),
    )
    audit.conditional_formatting.add(
        f"C4:C{audit.max_row}",
        FormulaRule(formula=['$C4="FAIL"'], fill=PatternFill("solid", fgColor="F4CCCC")),
    )

    # Number formatting for dashboards
    for ws in [entity_dash, combined_dash]:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.number_format = '$#,##0'
        for cell_ref in ["B8","B9","B10","B11","E6","E7","E8","E9","E10","E11"]:
            ws[cell_ref].number_format = '0.0%'

    desired_order = [
        "Cover & Instructions",
        "Assumptions & Methodology",
        "Entity Summary Dashboard",
        "Combined Portfolio Dashboard",
        "Investment Master",
        "Cash Flow Ledger",
        "Official Valuation Ledger",
        "Internal Valuation Model",
        "Exit Scenario Valuation",
        "Entity-Level Performance",
        "Investment-Level Performance",
        "Cash Flow Timeline",
        "Lookups & Data Validation",
        "Audit Checks",
    ]
    wb._sheets = [wb[name] for name in desired_order]

    wb.save(OUTPUT_FILE)


if __name__ == "__main__":
    os.makedirs(ROOT / "tools", exist_ok=True)
    create_workbook()
    print(OUTPUT_FILE)
